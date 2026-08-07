# ============================================================
# Scientific question / 科学问题:
# 综合名录匹配、坐标解析与地名地理编码结果，产出一份分类地位
# 完整、坐标可用、问题可追溯的两爬新纪录修订表。
# Assemble catalogue matching, coordinate parsing, and gazetteer
# geocoding into a revised, fully traceable herp-record table.
#
# Objective / 分析目标:
# 1) 填补/更正 科、属 中拉名并新增名录对应列（纲、目、拉丁名等）
# 2) 统一坐标为十进制度、更正互换与错位坐标、按小地名补坐标
# 3) 标记重复行、非两爬行、国外记录、年份/DOI 等问题
# 4) 输出 总表(修订) + 修订日志 + 审查问题清单 + 数据说明
#
# Input / 输入数据:
# - 两栖爬行动物数据合并表-8.7最新版.xlsx（原表，只读）
# - 动物界-脊索动物门-2026-10714.xlsx（名录）
# - match_result.pkl, coords.pkl（前序脚本输出）
# Output / 预期输出:
# - ~/Downloads/两栖爬行动物数据合并表-8.7修订完善版.xlsx
# Key assumptions / 关键假设:
# - 名录为分类地位权威；原 scientific_name 保留为发表所用名。
# - 仅删除 3 行确定的表头伪行；其余问题行保留并标记。
# Main packages / 主要包: pandas, openpyxl
# ============================================================

import pickle
import re
import unicodedata
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]  # 仓库根目录 / repo root
CAT_PATH = str(ROOT / "source_data/动物界-脊索动物门-2026-10714.xlsx")
REC_PATH = str(ROOT / "source_data/两栖爬行动物数据合并表-8.7最新版.xlsx")
OUT_PATH = str(ROOT / "source_data/两栖爬行动物数据合并表-8.7修订完善版.xlsx")
WORK = Path(__file__).parent / "intermediate"

# 确认为合并伪影的表头行（原表 Excel 行号）/ header artifact rows to drop
DROP_ROWS = {193, 1039, 2230}

# 地名地理编码结果（原表 Excel 行号 -> 坐标与说明）
# Gazetteer geocoding results keyed by original Excel row number
GEOCODE = {
    39:  (33.9917, 107.7984, "据小地名'小文公'定位太白山小文公庙（OSM），地点级"),
    162: (25.9800, 105.6133, "据小地名'滴水滩'定位关岭滴水滩瀑布（OSM），地点级"),
    167: (41.6118, 120.8335, "据小地名'黄半吉沟'定位北票市上园镇黄半吉沟（文献钻孔坐标），村级；注意该行为早白垩世化石记录"),
    302: (30.0980, 110.6100, "据小地名'后河村'定位五峰后河国家级自然保护区范围中心（110°30′-43′E,30°03′-09′N），保护区级"),
    358: (29.2742, 107.8416, "据原始文献：神仙洞位于重庆武隆江口镇（29°16.451′N,107°50.495′E，ZooKeys 2025），地点级；省份建议填'重庆'"),
    427: (25.6667, 107.8833, "原表仅有纬度；按原始描述（Gu et al. 2012 Zootaxa：Weng'ang, 25°40′N,107°53′E, 817 m）补经度，乡级"),
    743: (31.4810, 79.8020, "玛央河谷位于西藏札达县（Zhao 1998 模式产地），村级坐标未检索到，此为县驻地坐标，县级精度"),
    1343: (27.1183, 105.3246, "原始文献模式产地坐标：毕节七星关区五井保护区金甲寨村 27°7′5.92″N,105°19′28.47″E,1890 m（ZooKeys 2019）"),
    1345: (27.1183, 105.3246, "原始文献模式产地坐标：毕节七星关区五井保护区金甲寨村 27°7′5.92″N,105°19′28.47″E,1890 m（ZooKeys 2019）"),
    1390: (30.8002, 103.3840, "据'鸡冠山'定位崇州市文井江镇鸡冠山（OSM），乡级；该行另一地点兴沟村(天全)已发表坐标为 29.9303N,102.3882E"),
    2177: (22.3924, 114.0221, "'大揽'为香港大榄（棕脊蛇模式产地，文献为其地模标本记录），大榄郊野公园中心（OSM）；省份建议填'香港'"),
}
# 明确坐标错误的更正（原表行号 -> 正确坐标与依据）/ coordinate corrections
COORD_FIX = {
    1342: (27.6567, 105.3872, "原坐标(26.5,119.5)落在福建，与产地'青山村照子山保护区(贵州毕节)'不符；按原始文献模式坐标 27°39′24″N,105°23′14″E 更正"),
    1344: (27.6567, 105.3872, "原坐标(26.5,119.5)落在福建，与产地'青山村照子山保护区(贵州毕节)'不符；按原始文献模式坐标 27°39′24″N,105°23′14″E 更正"),
}
# 据引用格式可确定物种的行 / rows whose species is recoverable from citation
ROW_SPECIES_FIX = {
    1879: ("species_zh", "四爪陆龟", "该行物种字段全空，文献标准引用格式明确为'四爪陆龟'（新疆艾比湖保护区）"),
    1297: ("scientific_name", "Lepidodactylus lugubris", "物种列为空，引用格式为 Lepidodactylus lugubris（瓜德罗普外来记录）"),
    1299: ("scientific_name", "Lepidodactylus lugubris", "物种列为空，引用格式为 Lepidodactylus lugubris（瓜德罗普外来记录）"),
    1301: ("scientific_name", "Lepidodactylus lugubris", "物种列为空，引用格式为 Lepidodactylus lugubris（瓜德罗普外来记录）"),
}
# 非两爬/非动物类群行（原表行号 -> 类群说明）/ non-target-taxon rows
NON_TARGET = {
    28: "无学名的国外新种条目（斑蟾，中南美）", 29: "无学名的国外新种条目（玻璃蛙，南美）",
    30: "无学名的国外新种条目（玻璃蛙，南美）", 31: "无学名的国外新种条目（雨蛙）",
    32: "无学名的国外新种条目（毒蛙，南美）", 33: "鱼类（鲤），非两栖爬行动物",
    117: "淡水蟹（甲壳动物），非两栖爬行动物", 183: "蕨类植物（骨碎补科）",
    184: "水生植物", 186: "植物（水晶兰）", 340: "昆虫（松叶蜂）",
    494: "涡虫（扁形动物）", 830: "蜘蛛", 831: "蜘蛛",
    972: "软骨鱼（鳐），非两栖爬行动物", 1236: "水螨（蛛形纲）", 1237: "水螨（蛛形纲）",
    1238: "螯虾鳃腔共生扁虫/螨类记录", 1239: "水螨（蛛形纲）", 1240: "水螨（蛛形纲）",
    1241: "水螨（蛛形纲）", 1518: "蛙体内寄生线虫", 1519: "蛙体内寄生线虫",
    1575: "原尾虫（六足总纲）", 1633: "昆虫（隐翅虫）",
    211: "文本编码损坏（乱码行），建议溯源后重录或删除",
    1941: "仅有引用信息的空条目，建议溯源补齐或删除",
}
HERP_CLASSES = ("两栖纲", "爬行纲")

# 名录科拉丁名的科级拼法规范：名录将游蛇科写作亚科形式 Colubrinae，
# 写入 family_en 时统一为科级拼法（分类概念仍从名录）。
# Family-rank spelling fix for the catalogue's subfamily-form quirk.
FAMILY_EN_FIX = {"Colubrinae": "Colubridae"}


def norm(s):
    """全角转半角去空白；空值返回 None / normalize or None."""
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return None
    s = unicodedata.normalize("NFKC", str(s)).strip()
    return re.sub(r"\s+", " ", s) or None


def zh_primary(zh: str) -> str:
    """取名录中文名主名（去括号与斜杠别名）/ primary zh name."""
    base = re.sub(r"[（(].*?[)）]", "", zh)
    return re.split(r"[/、]", base)[0].strip() or zh


def main() -> None:
    cat = pd.read_excel(CAT_PATH)
    rec = pd.read_excel(REC_PATH, sheet_name="总表", dtype=str)
    with open(WORK / "match_result.pkl", "rb") as f:
        matches = pickle.load(f)
    with open(WORK / "coords.pkl", "rb") as f:
        coords = pickle.load(f)

    la_index = {norm(r["物种拉丁名"]): i for i, r in cat.iterrows() if norm(r["物种拉丁名"])}
    zh_index = {}
    for i, r in cat.iterrows():
        z = norm(r["物种中文名"])
        if z:
            zh_index.setdefault(zh_primary(z), i)

    revlog = []   # 修订日志 / change log
    issues = []   # 审查问题清单 / issue list

    def log(row_excel, col, old, new, reason):
        revlog.append({"原表行号": row_excel, "列": col,
                       "原值": "" if old is None else str(old),
                       "新值": "" if new is None else str(new), "依据": reason})

    def issue(row_excel, sp, typ, detail, advice=""):
        issues.append({"原表行号": row_excel, "物种": sp or "", "问题类型": typ,
                       "详情": detail, "处理建议": advice})

    # ---- 据引用补物种名并重新匹配 / citation-based species fixes ----
    for erow, (col, val, reason) in ROW_SPECIES_FIX.items():
        i = erow - 2
        old = rec.at[i, col]
        rec.at[i, col] = val
        log(erow, col, old, val, reason)
        hit = None
        if col == "scientific_name" and val in la_index:
            hit = la_index[val]
        elif col == "species_zh" and val in zh_index:
            hit = zh_index[val]
        if hit is not None:
            matches[i] = {"cat_idx": hit, "genus_idx": None,
                          "method": "据引用补名后匹配", "flags": ["物种名据文献引用补齐，请复核"]}

    n = len(rec)
    new_cols = {c: [None] * n for c in
                ["名录拉丁名", "名录中文名", "纲_名录", "目中文名_名录", "目拉丁名_名录",
                 "名录匹配方式", "名录匹配备注", "latitude_dd", "longitude_dd",
                 "坐标备注", "疑似重复组", "审查标记", "原表行号"]}

    # ---- 重复行分组 / duplicate grouping ----
    key = (rec["species_zh"].fillna("") + "|" + rec["scientific_name"].fillna("") + "|"
           + rec["locality_zh"].fillna("") + "|" + rec["文献标准引用格式"].fillna(""))
    grp_id, seen = {}, {}
    for i, k in key.items():
        if k.strip("|"):
            seen.setdefault(k, []).append(i)
    gn = 0
    for k, idxs in seen.items():
        if len(idxs) > 1:
            gn += 1
            for i in idxs:
                grp_id[i] = f"D{gn:03d}"

    # ---- 逐行处理 / per-row processing ----
    taxo_cols = [("family_en", "科拉丁名"), ("family_zh", "科中文名"),
                 ("genus_en", "属拉丁名"), ("genus_zh", "属中文名")]
    n_fill = n_corr = 0
    for i in range(n):
        erow = i + 2
        r = rec.iloc[i]
        m = matches[i]
        tags = []
        new_cols["原表行号"][i] = erow
        cat_row = None
        if m["cat_idx"] is not None:
            cat_row = cat.loc[m["cat_idx"]]
        elif m["genus_idx"] is not None:
            cat_row = cat.loc[m["genus_idx"]]

        # 分类地位填补与更正 / taxonomy fill & correction
        if cat_row is not None:
            for col, ccol in taxo_cols:
                want = norm(cat_row[ccol])
                if ccol == "科拉丁名" and want in FAMILY_EN_FIX:
                    want = FAMILY_EN_FIX[want]
                have = norm(r[col])
                if want is None:
                    continue
                if have is None:
                    rec.iat[i, rec.columns.get_loc(col)] = want
                    log(erow, col, None, want, "据名录填补空缺")
                    n_fill += 1
                elif have != want and re.sub(r"^Genus\s+", "", have) != want:
                    rec.iat[i, rec.columns.get_loc(col)] = want
                    log(erow, col, have, want, "与名录不一致，按名录更正（原值见日志）")
                    n_corr += 1
            if m["cat_idx"] is not None:
                new_cols["名录拉丁名"][i] = norm(cat_row["物种拉丁名"])
                new_cols["名录中文名"][i] = norm(cat_row["物种中文名"])
                # species_zh / scientific_name 空缺填补 / fill blank name cols
                if norm(r["species_zh"]) is None:
                    v = zh_primary(norm(cat_row["物种中文名"]) or "")
                    if v:
                        rec.iat[i, rec.columns.get_loc("species_zh")] = v
                        log(erow, "species_zh", None, v, "据名录中文名填补")
                        n_fill += 1
                if norm(r["scientific_name"]) is None:
                    v = norm(cat_row["物种拉丁名"])
                    rec.iat[i, rec.columns.get_loc("scientific_name")] = v
                    log(erow, "scientific_name", None, v, "据名录现行有效拉丁名填补")
                    n_fill += 1
            new_cols["纲_名录"][i] = cat_row["纲中文名"]
            new_cols["目中文名_名录"][i] = cat_row["目中文名"]
            new_cols["目拉丁名_名录"][i] = cat_row["目拉丁名"]
            if cat_row["纲中文名"] not in HERP_CLASSES:
                tags.append(f"名录归{cat_row['纲中文名']}，非两栖爬行动物")
                issue(erow, r["species_zh"] or r["scientific_name"], "非两爬（据名录）",
                      f"匹配到{cat_row['纲中文名']}/{cat_row['目中文名']}：{cat_row['物种中文名']} {cat_row['物种拉丁名']}",
                      "请核实该行是否保留在两爬表中")
        new_cols["名录匹配方式"][i] = m["method"]
        if m["flags"]:
            new_cols["名录匹配备注"][i] = "；".join(m["flags"])
            for fl in m["flags"]:
                if "请复核" in fl or "请人工确认" in fl or "请核实" in fl:
                    issue(erow, r["species_zh"] or r["scientific_name"], "物种匹配需复核", fl)
        if m["method"].startswith("未匹配") and erow not in NON_TARGET and erow not in DROP_ROWS:
            issue(erow, r["species_zh"] or r["scientific_name"], "名录未收录/未匹配",
                  f"{m['method']}：zh={r['species_zh']}, la={r['scientific_name']}",
                  "确认是否为国外类群、异名或名录未收录种")

        # 坐标 / coordinates
        c = coords[i]
        lat, lon = c["lat_dd"], c["lon_dd"]
        cnotes = list(c["issues"])
        if erow in COORD_FIX:
            flat, flon, why = COORD_FIX[erow]
            log(erow, "latitude", r["latitude"], f"{flat:.4f}°N", why)
            log(erow, "longitude", r["longitude"], f"{flon:.4f}°E", why)
            rec.iat[i, rec.columns.get_loc("latitude")] = f"{flat:.4f}°N"
            rec.iat[i, rec.columns.get_loc("longitude")] = f"{flon:.4f}°E"
            lat, lon = flat, flon
            cnotes = [x for x in cnotes if "省份" not in x] + ["坐标已按原始文献更正"]
        elif erow in GEOCODE:
            glat, glon, why = GEOCODE[erow]
            if norm(r["latitude"]) is None:
                rec.iat[i, rec.columns.get_loc("latitude")] = f"{glat:.4f}°N"
                log(erow, "latitude", None, f"{glat:.4f}°N", why)
            if norm(r["longitude"]) is None:
                rec.iat[i, rec.columns.get_loc("longitude")] = f"{glon:.4f}°E"
                log(erow, "longitude", None, f"{glon:.4f}°E", why)
            old_flag = r["经纬度是否为后续补充"]
            rec.iat[i, rec.columns.get_loc("经纬度是否为后续补充")] = "是"
            if norm(old_flag) != "是":
                log(erow, "经纬度是否为后续补充", old_flag, "是", "坐标为本次按小地名补充")
            lat = lat if lat is not None else glat
            lon = lon if lon is not None else glon
            cnotes.append(why)
        elif c["swapped"]:
            log(erow, "latitude", r["latitude"], f"{lat:.4f}°N", "经纬度互换错误，已对调")
            log(erow, "longitude", r["longitude"], f"{lon:.4f}°E", "经纬度互换错误，已对调")
            rec.iat[i, rec.columns.get_loc("latitude")] = f"{lat:.4f}°N"
            rec.iat[i, rec.columns.get_loc("longitude")] = f"{lon:.4f}°E"
        if lat is not None and abs(lat) <= 90:
            new_cols["latitude_dd"][i] = round(lat, 6)
        if lon is not None and abs(lon) <= 180:
            new_cols["longitude_dd"][i] = round(lon, 6)
        if cnotes:
            new_cols["坐标备注"][i] = "；".join(cnotes)
            for t in cnotes:
                if ("省份" in t) or ("超出" in t) or ("无法解析" in t) or ("互换" in t):
                    issue(erow, r["species_zh"] or r["scientific_name"], "坐标异常",
                          t + f"（lat={r['latitude']}, lon={r['longitude']}）", "请对照原文核实")

        # 其他标记 / other tags
        if erow in NON_TARGET:
            tags.append("非两爬/问题条目：" + NON_TARGET[erow])
            issue(erow, r["species_zh"] or r["scientific_name"], "非两爬或问题条目",
                  NON_TARGET[erow], "建议移出两爬表或补齐后保留")
        if i in grp_id:
            new_cols["疑似重复组"][i] = grp_id[i]
        se, sn = norm(r["species_en"]), norm(r["scientific_name"])
        if se and re.match(r"^[A-Z][a-z]+ [a-z]+$", se) and se != sn:
            tags.append("species_en 疑为拉丁名而非英文名")
        y1, y2 = norm(r["year"]), norm(r["发表年份"])
        if y1 and y2 and y1 != y2:
            tags.append(f"year({y1})与发表年份({y2})不一致")
            issue(erow, r["species_zh"] or r["scientific_name"], "年份不一致",
                  f"year={y1}，发表年份={y2}", "以原文为准统一")
        elif y1 and not y2:
            rec.iat[i, rec.columns.get_loc("发表年份")] = y1
            log(erow, "发表年份", None, y1, "据 year 列填补")
        elif y2 and not y1:
            rec.iat[i, rec.columns.get_loc("year")] = y2
            log(erow, "year", None, y2, "据 发表年份 列填补")

        doi = norm(r["DOI"])
        if doi:
            new_doi = re.sub(r"^(https?://(dx\.)?doi\.org/|doi[:：]\s*)", "", doi, flags=re.I)
            if new_doi != doi:
                rec.iat[i, rec.columns.get_loc("DOI")] = new_doi
                log(erow, "DOI", doi, new_doi, "DOI 统一为纯编号格式")
            if not new_doi.lower().startswith("10."):
                tags.append("DOI 列非标准 DOI（链接/其他文本）")

        rt = norm(r["record_type_zh"])
        if rt == "其他新记录":
            rec.iat[i, rec.columns.get_loc("record_type_zh")] = "其他新纪录"
            log(erow, "record_type_zh", rt, "其他新纪录", "用字统一（记录→纪录）")

        if tags:
            new_cols["审查标记"][i] = "；".join(tags)

    # ---- 拼装输出 / build output frame ----
    for cname, vals in new_cols.items():
        rec[cname] = vals
    drop_idx = [e - 2 for e in DROP_ROWS]
    for e in sorted(DROP_ROWS):
        r = rec.iloc[e - 2]
        log(e, "整行", f"{r['species_zh']}/{r['scientific_name']}", "已删除",
            "合并产生的表头伪行（species_zh/学名等列为字段名）")
    rec_out = rec.drop(index=drop_idx).reset_index(drop=True)

    # 重复组摘要 / duplicate summary into issues
    for k, idxs in seen.items():
        if len(idxs) > 1:
            rows = [x + 2 for x in idxs]
            sp = rec.iloc[idxs[0]]["species_zh"] or rec.iloc[idxs[0]]["scientific_name"]
            issue(rows[0], sp, "疑似重复行",
                  f"{len(idxs)}行内容（物种+地点+引用）相同：原表行 {rows}（组{grp_id[idxs[0]]}）",
                  "建议保留信息最全的一行，其余删除")

    # ---- 写工作簿 / write workbook ----
    wb = Workbook()
    ws = wb.active
    ws.title = "总表"
    header_font = Font(bold=True)
    new_fill = PatternFill("solid", fgColor="FFF2CC")  # 新增列表头底色 / new-col header fill
    cols = list(rec_out.columns)
    ws.append(cols)
    n_orig_cols = 32
    for j, cname in enumerate(cols, 1):
        cell = ws.cell(row=1, column=j)
        cell.font = header_font
        if j > n_orig_cols:
            cell.fill = new_fill
    for _, row in rec_out.iterrows():
        ws.append([None if (v is None or (isinstance(v, float) and pd.isna(v)) or v == "")
                   else v for v in row.tolist()])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(rec_out) + 1}"
    for j, cname in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(j)].width = max(10, min(28, len(str(cname)) * 2 + 6))

    def add_sheet(name, df):
        s = wb.create_sheet(name)
        s.append(list(df.columns))
        for j in range(1, len(df.columns) + 1):
            s.cell(row=1, column=j).font = header_font
        for _, row in df.iterrows():
            s.append(list(row))
        s.freeze_panes = "A2"
        widths = {"依据": 60, "详情": 60, "处理建议": 30, "原值": 24, "新值": 24, "问题类型": 16}
        for j, cn in enumerate(df.columns, 1):
            s.column_dimensions[get_column_letter(j)].width = widths.get(cn, 14)
        return s

    df_log = pd.DataFrame(revlog, columns=["原表行号", "列", "原值", "新值", "依据"]).sort_values(
        ["原表行号", "列"]).reset_index(drop=True)
    add_sheet("修订日志", df_log)
    order = {"非两爬（据名录）": 0, "非两爬或问题条目": 0, "坐标异常": 1, "疑似重复行": 2,
             "物种匹配需复核": 3, "名录未收录/未匹配": 4, "年份不一致": 5}
    df_iss = pd.DataFrame(issues, columns=["原表行号", "物种", "问题类型", "详情", "处理建议"])
    df_iss["_o"] = df_iss["问题类型"].map(order).fillna(9)
    df_iss = df_iss.sort_values(["_o", "原表行号"]).drop(columns="_o").reset_index(drop=True)
    add_sheet("审查问题清单", df_iss)

    stat = pd.Series([m["method"] for m in matches]).value_counts()
    notes = [
        ["两栖爬行动物新纪录数据合并表 —— 修订完善版说明", ""],
        ["生成日期", "2026-08-07"],
        ["原始文件", "两栖爬行动物数据合并表-8.7最新版.xlsx（未改动）"],
        ["分类标准", "《中国生物物种名录》2026版 动物界-脊索动物门（10714种）"],
        ["", ""],
        ["一、分类地位处理", ""],
        ["匹配策略", "拉丁名精确→中文名精确(含名录括号/斜杠别名)→三名法亚种归并→拉丁名空格/粘连修复→中文亚种归并→人工核定异名表→种加词词干推断(带目/科相容与原属存在性护栏)→属级→中文名模糊(换/增删一字，需种加词不矛盾)"],
        ["匹配结果", "；".join(f"{k}:{v}行" for k, v in stat.items())],
        ["填补空缺单元格数", f"{n_fill}（science/family/genus 等列按名录补齐）"],
        ["更正与名录不符单元格数", f"{n_corr}（原值全部保留在修订日志中）"],
        ["新增列", "名录拉丁名/名录中文名（名录现行有效名）；纲_名录、目中文名_名录、目拉丁名_名录；名录匹配方式与备注"],
        ["重要说明", "scientific_name 保留发表时所用名；'名录拉丁名'为现行有效名，两者不同即提示异名关系。推断类匹配均带'请复核'标记。"],
        ["科名拼法", "名录将游蛇科拉丁名写作亚科形式 Colubrinae；本表 family_en 统一采用科级拼法 Colubridae（分类概念仍按名录，即游蛇科含 Natricinae 等）。"],
        ["", ""],
        ["二、坐标处理", ""],
        ["十进制列", "latitude_dd/longitude_dd 由原文本解析（度分、度分秒、全角字符均已处理；范围值取中点并注明）"],
        ["互换更正", "5行纬度>90°且与经度对调后合理，已对调（行1211,1214,1222,1229,1432）"],
        ["错位更正", "行1342,1344 坐标与产地(贵州毕节)不符，按原始文献模式产地坐标更正"],
        ["小地名补坐标", f"{len(GEOCODE)}行（小文公/滴水滩/黄半吉沟/后河村/神仙洞/玛央河谷/金甲寨村/鸡冠山/大揽/翁昂），来源与精度见'坐标备注'列；'经纬度是否为后续补充'已置'是'"],
        ["未补坐标", "仅有省级信息的141行未按省中心填坐标（避免虚假精度）；'三楼村''七里地区'等无法唯一定位的地名未填并已标记"],
        ["", ""],
        ["三、其他审查", ""],
        ["删除行", "3行合并产生的表头伪行（原表行193,1039,2230），见修订日志"],
        ["疑似重复", "同物种+地点+引用的重复行已按组编号（疑似重复组列），建议每组保留信息最全的一行"],
        ["非两爬条目", "经名录及人工判别，蝙蝠/鸟/鱼/蟹/水螨/线虫/蜘蛛/昆虫/植物等混入条目已在'审查标记'列标注"],
        ["年份与DOI", "year与发表年份互补填齐，冲突152行已标记；DOI统一为纯编号，非DOI链接已标注"],
        ["用字统一", "record_type_zh '其他新记录'→'其他新纪录'"],
        ["", ""],
        ["注意（茂兰瘰螈 行427）", "原表仅有纬度25.6667°N，与原始描述25°40′N一致；经度按原始描述107°53′E补为107.8833°E"],
    ]
    s = wb.create_sheet("数据说明")
    for row in notes:
        s.append(row)
    s.column_dimensions["A"].width = 28
    s.column_dimensions["B"].width = 130
    for cell in s["A"]:
        cell.font = Font(bold=True)
    s["A1"].font = Font(bold=True, size=14)
    for row in s.iter_rows(min_col=2, max_col=2):
        row[0].alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(OUT_PATH)
    print(f"已输出: {OUT_PATH}")
    print(f"总表行数: {len(rec_out)}（原 {n} 行，删除 {len(DROP_ROWS)} 行表头伪行）")
    print(f"修订日志条目: {len(df_log)}；问题清单条目: {len(df_iss)}")
    print(f"填补: {n_fill} 格；更正: {n_corr} 格")


if __name__ == "__main__":
    main()
